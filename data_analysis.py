

import openpyxl


#replicates
r = 3
#samples, the row is shifted by 13 in the excel output
m = 111
welln = m + 13
#define the names for the conditions and DNA. this will be input on the interface at the 1st step.

CONDNAMES = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
DNANAMES = ['ThermoPrep', 'Midi', 'PrecMidi']

wb = openpyxl.load_workbook('data.xlsx')
sheets = wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Range 1')

#create the names
names = []
name = ''
for i in DNANAMES:
    name = i
    for j in CONDNAMES:
        name = name + ' / ' + j
        names.append(name)
        name = i

#read in the dataand calculate the basal level from blanks
rawdata = []
for i in range(13, welln):
       rawdata.append((sheet.cell(row=i, column=4).value))
blanks = []
x = 0
for i in range(r):
    blanks.append(rawdata[len(names)*r + x])
    x = x+1
blankaverage = (sum(blanks)/r)
print (blanks)
print (blankaverage)

#do the blank substraction
data = []
for i in range(13, welln):
        data.append( (sheet.cell(row=i, column=4).value - blankaverage))

#print out the reorganised data
procdata = open('procdata.csv','w')
x = 0
for i in range(len(names)):
    procdata.write('\n')
    procdata.write (names[i])
    for j in range(r):
        procdata.write (' ,' + str(data[x]))
        x = x+1
